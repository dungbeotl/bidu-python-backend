from typing import List, Dict, Any
from datetime import datetime
import json
import io
import csv
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class ExportUtil:
    def __init__(self):
        pass

    async def _export_dataset_to_json(
        self, data: List[Dict[str, Any]]
    ) -> StreamingResponse:
        """
        Xuất dữ liệu thành JSON (mỗi record một dòng).

        Args:
            data: Danh sách dữ liệu đã xử lý.

        Returns:
            StreamingResponse với dữ liệu JSON.
        """
        # AWS Personalize yêu cầu mỗi record trên một dòng không có dấu phẩy ở cuối
        jsonl_output = ""
        for item in data:
            jsonl_output += json.dumps(item) + "\n"

        # Trả về response
        return StreamingResponse(
            iter([jsonl_output]),
            media_type="application/json",
            headers={
                "Content-Disposition": f"attachment; filename=products_dataset_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jsonl"
            },
        )

    async def _export_dataset_to_csv(
        self, data: List[Dict[str, Any]]
    ) -> StreamingResponse:
        """
        Xuất dữ liệu thành CSV.

        Args:
            data: Danh sách dữ liệu đã xử lý.

        Returns:
            StreamingResponse với dữ liệu CSV.
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Viết header
        if data:
            writer.writerow(data[0].keys())

        # Viết dữ liệu
        for item in data:
            writer.writerow(item.values())

        # Reset về đầu file
        output.seek(0)

        # Trả về response
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename=products_dataset_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            },
        )

    async def _export_dataset_to_excel(
        self, data: List[Dict[str, Any]], filename_prefix: str = "data"
    ) -> StreamingResponse:
        """
        Xuất dữ liệu thành file Excel.

        Args:
            data: Danh sách dữ liệu đã xử lý.
            filename_prefix: Tiền tố cho tên file.

        Returns:
            StreamingResponse với dữ liệu Excel.
        """
        # Tạo workbook và worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        if not data:
            # Nếu không có dữ liệu, tạo file trống
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return StreamingResponse(
                io.BytesIO(output.read()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename={filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                },
            )

        # Lấy headers từ record đầu tiên
        headers = list(data[0].keys())
        
        # Thêm header với style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Thêm dữ liệu
        for row_num, item in enumerate(data, 2):
            for col_num, header in enumerate(headers, 1):
                value = item.get(header, "")
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Tự động điều chỉnh độ rộng cột
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Giới hạn tối đa 50 ký tự
            ws.column_dimensions[column_letter].width = adjusted_width

        # Lưu vào BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            },
        )

    async def _export_multiple_sheets_to_excel(
        self, 
        sheets_data: Dict[str, List[Dict[str, Any]]], 
        filename_prefix: str = "multi_sheet_data"
    ) -> StreamingResponse:
        """
        Xuất dữ liệu thành file Excel với nhiều sheet.

        Args:
            sheets_data: Dictionary với key là tên sheet và value là dữ liệu cho sheet đó.
            filename_prefix: Tiền tố cho tên file.

        Returns:
            StreamingResponse với dữ liệu Excel.
        """
        # Tạo workbook
        wb = Workbook()
        
        # Xóa sheet mặc định
        wb.remove(wb.active)

        # Kiểm tra nếu không có dữ liệu
        if not sheets_data:
            # Tạo một sheet trống
            ws = wb.create_sheet("Empty")
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return StreamingResponse(
                io.BytesIO(output.read()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename={filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                },
            )

        # Tạo từng sheet
        for sheet_name, data in sheets_data.items():
            # Tạo sheet mới
            ws = wb.create_sheet(title=sheet_name)
            
            if not data:
                # Nếu sheet không có dữ liệu, bỏ qua
                continue
            
            # Lấy headers từ record đầu tiên
            headers = list(data[0].keys())
            
            # Thêm header với style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Thêm dữ liệu
            for row_num, item in enumerate(data, 2):
                for col_num, header in enumerate(headers, 1):
                    value = item.get(header, "")
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            # Tự động điều chỉnh độ rộng cột
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Giới hạn tối đa 50 ký tự
                ws.column_dimensions[column_letter].width = adjusted_width

        # Lưu vào BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            },
        )
